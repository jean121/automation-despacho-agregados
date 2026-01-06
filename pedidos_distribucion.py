# -*- coding: utf-8 -*-
"""
Automatización de UNICON (PEDIDOS_DISTRIBUCION - AGREGADOS) desde Citrix
============================================================================

Este script lee pedidos desde una tabla en Excel y los ingresa en la ventana
remota (Citrix) usando teclas (send keys), TABs y selección por imagen/OCR.

⚠️ Requisitos (instalar en tu PC Windows donde correrá el script):
    pip install openpyxl pyautogui pillow pytesseract

Opcional (para mejor foco de ventana):
    pip install pygetwindow

Notas importantes:
- Cierra o minimiza ventanas innecesarias. Deja VS Code/terminal adelante y la
  ventana de UNICON inmediatamente detrás (Alt+Tab debe llevarte a UNICON).
- Prepara una imagen de referencia del botón "Salidas" tomada de tu pantalla
  y guárdala como 'salidas.png' en la misma carpeta del script. Puedes ajustar
  el nombre en SALIDAS_IMG_PATH.
- Ajusta los delays si ves que el Citrix/UNICON responde lento.
- Si quieres validar el flujo sin enviar teclas, usa DRY_RUN=True.

Autor: Elmer Jean Pierre Alpiste Ramirez
Archivo destino: C:\\Users\\ealpiste\\Documents\\Mis scripts\\pedidos_distribucion.py
"""

import time
import sys
import os
import tempfile
import shutil
from datetime import datetime
from pywinauto.keyboard import send_keys
from typing import List, Tuple
import warnings

# === Librerías de Excel ===
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# === Automatización por teclado/imagen ===
try:
    import pyautogui
except Exception as e:
    print("[ERROR] pyautogui no está disponible. Instálalo con: pip install pyautogui")
    raise

try:
    from PIL import Image
except Exception:
    Image = None

# OCR opcional
try:
    import pytesseract
except Exception:
    pytesseract = None

# Opcional para foco por título de ventana
try:
    import pygetwindow as gw
except Exception:
    gw = None

# ====== VARIABLES AJUSTABLES (CAMBIA AQUÍ) ======
EXCEL_PATH = r"C:\Users\ealpiste\OneDrive - Unacem.corp\Compartido Victor\DESPACHO DE AGREGADOS_YB 2025 2.3.xlsx"
TABLE_NAME = "Tabla27182"            # Nombre de la tabla (variable)
START_ROW_IN_TABLE = 4               # Fila inicial dentro de la tabla (sin contar cabecera)
END_ROW_IN_TABLE =   14                 # Fila final dentro de la tabla (sin contar cabecera)
TARGET_COLUMN_INDEX1 = 7              # 7ma columna de la tabla (agregado-destino)
TARGET_COLUMN_INDEX2 = 6              # 6ta columna de la tabla (cubicaje)

# Ventana (solo informativo; si instalas pygetwindow, puedes usarlo para enfocar)
WINDOW_TITLE_HINT = "UNICON - Módulo de PEDIDOS_DISTRIBUCION - AGREGADOS"

# Imagen del botón "Salidas" (captura desde tu pantalla)
SALIDAS_IMG_PATH = "salidas.png"
SALIDAS_IMG_CONFIDENCE = 0.85  # baja si la UI tiene leves cambios

# Delays / comportamiento
DELAY_SHORT = 0.10   # pequeño entre teclas
DELAY_MED = 0.25     # mediano entre pasos
DELAY_LONG = 0.60    # largo cuando la UI cambia de grilla
WAIT_AFTER_REFRESH = 1.20  # después de presionar 'b'

DRY_RUN = False  # True para simular sin enviar teclas

# ====== MAPEOS DE NAVEGACIÓN (ajusta si cambia el orden en tu UI) ======
PLANTA_TO_DOWN_PRESSES = {
    # MEIGGS -> 0 (ya está seleccionada por defecto)
    "MEIGGS": 1,
    "OQUENDO": 0,
    "MATERIALES": 3,
    "COLLIQUE": 2,
}

AGREGADO_TO_DOWN_PRESSES = {
    # 5 -> 0 (ya seleccionado)
    "5": 0,
    "AR": 1,     # Arena
    "67": 2,     # Piedra Huso 67
    "89": 3,     # Piedra Huso 89
}

# ======================================================================
# Utilidades
# ======================================================================
def log(msg: str) -> None:
    print(msg)
    sys.stdout.flush()

def hotkey(*keys: str):
    """Envía una combinación de hotkeys."""
    if DRY_RUN:
        log(f"[DRY] hotkey{keys}")
        return
    pyautogui.hotkey(*keys)

def type_text(text: str):
    if DRY_RUN:
        log(f"[DRY] type_text('{text}')")
        return
    try:
        # pywinauto.send_keys suele funcionar mejor en sesiones remotas/Citrix.
        # send_keys interpreta modificadores especiales, pero para números/puntos no hay problema.
        send_keys(str(text), pause=0.02)
    except Exception:
        # Fallback a pyautogui si send_keys falla por alguna razón.
        pyautogui.typewrite(str(text), interval=0.02)

def focus_unicon_window():
    """Intenta enfocar la ventana de UNICON.
    - Si pygetwindow está disponible, busca por título.
    - Si no, usa Alt+Tab suponiendo que UNICON está inmediatamente detrás.
    """
    """ if gw is not None:
        try:
            windows = [w for w in gw.getAllTitles() if WINDOW_TITLE_HINT.lower() in w.lower()]
            if windows:
                w = gw.getWindowsWithTitle(windows[0])[0]
                w.activate()
                time.sleep(DELAY_MED)
                return True
        except Exception:
            pass """
    # Fallback Alt+Tab
    hotkey('alt', 'tab')
    time.sleep(DELAY_MED)
    print("[INFO] Enfocamos la ventana PDIDOS_DISTRIBUCIÓN usando Alt+Tab (asegúrate que esté justo detrás).")
    return True

def locate_and_click_salidas() -> bool:
    """Intenta hacer clic en el botón 'Salidas' por imagen u OCR.
    Retorna True si se hizo clic, False si no.
    """
    # Imagen
    if os.path.exists(SALIDAS_IMG_PATH):
        try:
            if DRY_RUN:
                log("[DRY] Buscar y clic en imagen 'Salidas'")
                return True
            box = pyautogui.locateOnScreen(SALIDAS_IMG_PATH, confidence=SALIDAS_IMG_CONFIDENCE)
            if box:
                center = pyautogui.center(box)
                pyautogui.moveTo(center.x, center.y, duration=0.1)
                pyautogui.click()
                time.sleep(DELAY_MED)
                return True
        except Exception:
            pass
    # OCR (si disponible): buscar texto 'Salidas' en pantalla
    if pytesseract is not None and Image is not None:
        try:
            screenshot = pyautogui.screenshot()
            text = pytesseract.image_to_string(screenshot)
            if 'Salidas' in text:
                # Si no tenemos coords exactas, intenta navegar con TAB hasta encontrar foco
                # Aquí hacemos 2 intentos de TAB + SPACE por si el foco cae en el botón.
                for _ in range(12):
                    if DRY_RUN:
                        log("[DRY] send_keys('{TAB}')")
                    else:
                        send_keys('{TAB}', pause=DELAY_SHORT)
                    time.sleep(DELAY_SHORT)
                    if DRY_RUN:
                        log("[DRY] send_keys('{SPACE}')")
                    else:
                        send_keys('{SPACE}', pause=DELAY_SHORT)
                    time.sleep(DELAY_SHORT)
                return True
        except Exception:
            pass
    log("[WARN] No se pudo localizar el botón 'Salidas'. Asegúrate de preparar 'salidas.png' o ajustar el flujo de TABs.")
    return False

# ====== UTILIDADES ADICIONALES ======
def crear_copia_temporal(xlsx_path: str) -> str:
    """Crea una copia temporal del archivo xlsx y devuelve la ruta (delete=False)."""
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    shutil.copy2(xlsx_path, tmp_path)
    return tmp_path

# ======================================================================
# Lectura del Excel (tabla)
# ======================================================================
def leer_pedidos_desde_excel() -> List[Tuple[str, str, float]]:
    """Lee (agregado-destino, planta, cubicaje) por fila desde la tabla.

    Retorna una lista de tuplas: (agregado, planta, cubicaje)
    donde 'agregado' es uno de {'5','AR','67','89'} y 'planta' es
    {'MEIGGS','OQUENDO','MATERIALES','COLLIQUE'}.
    """
    # Hoja del día y mes actual en formato dd.mm
    #sheet_name = datetime.now().strftime('%d.%m')
    sheet_name = "07.01"  # funcionará siempre que la hoja con ese nombre exista exactamente

    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"No existe el archivo de Excel: {EXCEL_PATH}")

    wb = None
    tmp_path = None
    try:
        try:
            # Suprimir la advertencia conocida de openpyxl al leer ciertas validaciones
            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore",
                    message=".*Data Validation extension is not supported.*",
                    category=UserWarning,
                )
                wb = load_workbook(EXCEL_PATH, data_only=True)
        except (PermissionError, OSError) as e:
            # Intentar crear y usar una copia temporal si el archivo original está bloqueado
            log(f"[WARN] No se pudo abrir '{EXCEL_PATH}' directamente: {e}. Intentando copia temporal...")
            try:
                tmp_path = crear_copia_temporal(EXCEL_PATH)
                with warnings.catch_warnings():
                    warnings.filterwarnings(
                        "ignore",
                        message=".*Data Validation extension is not supported.*",
                        category=UserWarning,
                    )
                    wb = load_workbook(tmp_path, data_only=True)
            except Exception:
                # Re-raise original error si no se puede leer ni la copia
                raise
    finally:
        # eliminar la copia temporal si fue creada
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass

    if wb is None:
        raise ValueError("No se pudo cargar el libro de Excel.")

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{sheet_name}' no existe en el Excel. Verifica el nombre (formato dd.mm).")

    ws = wb[sheet_name]

    # Obtener la tabla por nombre
    table_obj = None
    try:
        # openpyxl >=3.1
        if hasattr(ws, 'tables') and isinstance(ws.tables, dict) and TABLE_NAME in ws.tables:
            table_obj = ws.tables[TABLE_NAME]
        else:
            # Compatibilidad con versiones que exponen _tables como lista
            for t in getattr(ws, '_tables', []):
                if getattr(t, 'name', None) == TABLE_NAME:
                    table_obj = t
                    break
    except Exception:
        pass

    if table_obj is None:
        raise ValueError(f"No se encontró la tabla '{TABLE_NAME}' en la hoja '{sheet_name}'.")

    # Rango de la tabla, ej. 'A1:M200'
    min_col, min_row, max_col, max_row = range_boundaries(table_obj.ref)

    # Columnas destino dentro del rango de la tabla
    col_agregado_destino = min_col + (TARGET_COLUMN_INDEX1 - 1)
    col_cubicaje = min_col + (TARGET_COLUMN_INDEX2 - 1)

    # Filas reales dentro del rango (sumar header: +1)
    start_row = min_row + START_ROW_IN_TABLE
    end_row = min_row + END_ROW_IN_TABLE

    pedidos: List[Tuple[str, str, float]] = []

    for r in range(start_row, end_row + 1):
        celda_ag_dest = ws.cell(row=r, column=col_agregado_destino).value
        celda_cubicaje = ws.cell(row=r, column=col_cubicaje).value

        if celda_ag_dest is None:
            continue

        texto = str(celda_ag_dest).strip()
        # esperado: 'AR-MEIGGS', '67-MATERIALES', etc.
        if '-' not in texto:
            log(f"[WARN] Formato inesperado en fila {r}: '{texto}' (se espera 'AGREGADO-DESTINO')")
            continue
        agregado, planta = [s.strip().upper() for s in texto.split('-', 1)]

        # normalizar agregado numérico a string
        if agregado in {'5', '67', '89', 'AR'}:
            pass
        else:
            # intentar quitar sufijos (p.ej. '5 ')
            agregado = agregado.replace('PIEDRA', '').replace('HUSO', '').strip()

        # cubicaje numérico
        try:
            cubicaje = float(str(celda_cubicaje).replace(',', '').strip())
        except Exception:
            log(f"[WARN] Cubicaje inválido en fila {r}: '{celda_cubicaje}'")
            continue

        pedidos.append((agregado, planta, cubicaje))

    return pedidos

# ======================================================================
# Flujo principal de envío a UNICON
# ======================================================================
def procesar_pedido(index: int, agregado: str, planta: str, cubicaje: float) -> None:
    log(f"\n[INFO] Procesando pedido: {index} | Agregado='{agregado}' | Planta='{planta}' | Cubicaje={cubicaje}")

    

    # 2) refrescar búsqueda (posicionamiento inicial)
    if DRY_RUN:
        log("[DRY] send_keys('b')")
    else:
        send_keys('b', pause=DELAY_SHORT)
    time.sleep(WAIT_AFTER_REFRESH)

    # 3) 2 TABs para ir a primera fila de la grilla superior (selección de planta)
    send_keys("{TAB}{TAB}")
    time.sleep(DELAY_MED)

    # 4) Seleccionar planta por flechas abajo
    down_planta = PLANTA_TO_DOWN_PRESSES.get(planta.upper(), 0)
    if down_planta > 0:
        # Usar send_keys para flechas
        if DRY_RUN:
            log(f"[DRY] send_keys('{{DOWN}}' * {down_planta})")
        else:
            send_keys("{DOWN}" * down_planta, pause=DELAY_SHORT)
        time.sleep(DELAY_MED)

    # 5) Pasar a la grilla inferior: tab, luego flecha abajo, clic en "Salidas", luego 8 tab
    send_keys("{TAB}")
    time.sleep(DELAY_SHORT)
    # reemplazo de press('down')
    if DRY_RUN:
        log("[DRY] send_keys('{DOWN}')")
    else:
        send_keys('{DOWN}', pause=DELAY_SHORT)
    time.sleep(DELAY_SHORT)
    locate_and_click_salidas()
    time.sleep(DELAY_MED)
    # reemplazo de press('tab', presses=8)
    if DRY_RUN:
        log("[DRY] send_keys('{TAB}' * 8)")
    else:
        send_keys('{TAB}' * 8, pause=DELAY_SHORT)
    time.sleep(DELAY_LONG)

    # 6) Seleccionar agregado en segunda tablilla
    down_agregado = AGREGADO_TO_DOWN_PRESSES.get(agregado.upper(), 0)
    if down_agregado > 0:
        if DRY_RUN:
            log(f"[DRY] send_keys('{{DOWN}}' * {down_agregado})")
        else:
            send_keys('{DOWN}' * down_agregado, pause=DELAY_SHORT)
        time.sleep(DELAY_MED)

    # 7) Seleccionar el número por defecto y sobreescribir con cubicaje
    # usar pywinauto.send_keys con modificadores: ^ = Ctrl, + = Shift, {RIGHT} = flecha derecha
    if DRY_RUN:
        log("[DRY] send_keys('^+{RIGHT}')")
    else:
        send_keys('^+{RIGHT}', pause=DELAY_SHORT)
    time.sleep(DELAY_SHORT)
    type_text(str(int(cubicaje) if cubicaje.is_integer() else cubicaje))
    time.sleep(DELAY_SHORT)

    # 8) Tab -> check -> despachar -> confirmar -> aceptar aviso
    # reemplazo de press('tab')
    if DRY_RUN:
        log("[DRY] send_keys('{TAB}')")
    else:
        send_keys('{TAB}', pause=DELAY_SHORT)
    time.sleep(DELAY_SHORT)
    # reemplazo de press('space')
    if DRY_RUN:
        log("[DRY] send_keys('{SPACE}')")
    else:
        send_keys('{SPACE}', pause=DELAY_SHORT)
    time.sleep(DELAY_SHORT)
    # reemplazo de press('d')
    if DRY_RUN:
        log("[DRY] send_keys('d')")
    else:
        send_keys('d', pause=DELAY_SHORT)
    time.sleep(DELAY_MED)
    # reemplazo de press('y')
    if DRY_RUN:
        log("[DRY] send_keys('y')")
    else:
        send_keys('y', pause=DELAY_MED)
    time.sleep(DELAY_MED)
    # reemplazo de press('space')
    if DRY_RUN:
        log("[DRY] send_keys('{SPACE}')")
    else:
        send_keys('{SPACE}', pause=DELAY_MED)
    time.sleep(DELAY_MED)

    # 9) 1? TAB para volver al punto inicial
    if DRY_RUN:
        log("[DRY] send_keys('{TAB}' * 4)")
    else:
        send_keys('{TAB}', pause=DELAY_MED)
    time.sleep(DELAY_MED)

    log(f"[OK] Pedido {index} procesado.")

def main():
    # Leer pedidos
    pedidos = leer_pedidos_desde_excel()
    if not pedidos:
        log("[WARN] No se encontraron pedidos válidos en el rango especificado.")
        return
    
    # 1) Enfocar UNICON
    focus_unicon_window()

    # Iterar por fila -> un pedido por fila
    for i, (agregado, planta, cubicaje) in enumerate(pedidos, start=1):
        procesar_pedido(i, agregado, planta, cubicaje)
        # Pequeña pausa entre pedidos por estabilidad
        time.sleep(0.8)

    log("\n[DONE] Se procesaron todos los pedidos del rango indicado.")

if __name__ == '__main__':
    main()
