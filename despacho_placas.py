# -*- coding: utf-8 -*-
"""
Automatización de despacho UNICON leyendo placas desde una tabla de Excel.

Autor: Elmer-ready
Notas:
- El script NO depende de la ventana de Excel: lee el archivo directamente con openpyxl.
- Asegúrate de que la ventana remota de UNICON esté al frente y en el estado inicial correcto.
- Por defecto espera F8 para continuar tras la selección manual del conductor. Si 'keyboard' no está instalado, pedirá Enter.

"""

import time
import sys
from typing import List, Optional
import warnings
from datetime import date
import ctypes
from pywinauto.keyboard import send_keys
import shutil
import tempfile
import os

# ====== VARIABLES AJUSTABLES (CAMBIA AQUÍ) ======
EXCEL_PATH = r"C:\Users\ealpiste\OneDrive - Unacem.corp\Compartido Victor\DESPACHO DE AGREGADOS_YB 2025 2.3.xlsx"  # Cambia a tu ruta real
TABLE_NAME = "Tabla276"            # Nombre de la tabla (variable)
START_ROW_IN_TABLE = 38             # Fila inicial dentro de la tabla (sin contar cabecera)
END_ROW_IN_TABLE =   40                # Fila final dentro de la tabla (sin contar cabecera)
TARGET_COLUMN_INDEX = 3              # 3ª columna de la tabla (1 = primera, 2 = segunda, 3 = tercera)

# Ventana remota (referencial, no usada por pyautogui directamente; sirve como documentación)
WINDOW_TITLE_REMOTO = r"UNICON  - Módulo de ALMACEN - ELMER JEAN PIERRE ALPISTE RAMIRE - \\Remota"

# Parámetros de navegación
SHIFT_TABS_A_BOTON_NOMBRE = 8       # Cantidad de Shift+Tab para llegar al botón sin nombre
FILTRO_NOMBRE_TEXTO = "alp"         # Texto del filtro para seleccionar "Alpiste Ramírez"
KEY_CONTINUAR = "f8"                 # Tecla que el usuario presionará para continuar tras seleccionar conductor
DELAY_CORTO = 0.01                    # Pequeñas esperas entre teclas
DELAY_MEDIO = 0.25
DELAY_LARGO = 0.6

# ====== IMPORTS PARA EXCEL Y TECLADO ======
try:
    import openpyxl
    from openpyxl.utils.cell import get_column_letter
    from openpyxl.utils import range_boundaries
except Exception as e:
    print("ERROR: No se pudo importar openpyxl. Instálalo con: pip install openpyxl")
    raise

try:
    import pyautogui as pag
    pag.FAILSAFE = True  # Esquina sup. izq para abortar
except Exception as e:
    print("ERROR: No se pudo importar pyautogui. Instálalo con: pip install pyautogui")
    raise

try:
    import pyperclip
except Exception as e:
    print("ERROR: No se pudo importar pyperclip. Instálalo con: pip install pyperclip")
    raise

# 'keyboard' es opcional: si no está, se usa input() como alternativa
try:
    import keyboard as kb
    KEYBOARD_AVAILABLE = True
except Exception:
    KEYBOARD_AVAILABLE = False
# ====== IMPORTS PARA CONEXIÓN VENTANA REMOTA ======
try:
    from pywinauto import Desktop, Application
except Exception:
    print("ERROR: No se pudo importar pywinauto. Instálalo con: pip install pywinauto")
    raise


# ====== UTILIDADES EXCEL ======
def crear_copia_temporal(xlsx_path: str) -> str:
    """Crea una copia temporal del archivo xlsx y devuelve la ruta (delete=False)."""
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    shutil.copy2(xlsx_path, tmp_path)
    return tmp_path


def encontrar_tabla_en_libro(xlsx_path: str, table_name: str):
    """
    Busca la tabla por nombre en todas las hojas del libro y devuelve
    (wb, ws, ref, min_col, min_row, max_col, max_row, temp_path).
    temp_path es None si no se creó copia temporal, o la ruta al temporal si sí.
    """
    temp_path = None
    # Suprimir warnings conocidos de openpyxl sobre Data Validation durante la carga
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Data Validation extension is not supported")
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except (PermissionError, OSError) as e:
            print(f"[WARN] No se pudo abrir '{xlsx_path}' directamente ({e}). Creando copia temporal...")
            temp_path = crear_copia_temporal(xlsx_path)
            print(f"[INFO] Copia temporal creada")
            wb = openpyxl.load_workbook(temp_path, data_only=True)

    # Intentar seleccionar directamente la hoja con nombre "D.M" o "DD.MM" (p.ej. 29.12)
    today = date.today()
    candidates_names = {f"{today.day}.{today.month}", f"{today.day:02d}.{today.month:02d}"}
    ws_list = []
    for name in candidates_names:
        if name in wb.sheetnames:
            ws_list = [wb[name]]
            break
    # Si no encontramos la hoja con el formato día.mes, procesar todas las hojas (fallback)
    if not ws_list:
        ws_list = wb.worksheets

    for ws in ws_list:
        tablas = {}
        if hasattr(ws, "_tables") and isinstance(ws._tables, dict):
            tablas.update(ws._tables)
        if hasattr(ws, "tables"):
            try:
                tablas.update(ws.tables)
            except Exception:
                pass

        for nombre, tbl in tablas.items():
            nombre_tbl = getattr(tbl, "name", None) or getattr(tbl, "displayName", None) or nombre
            if nombre_tbl == table_name:
                ref = getattr(tbl, "ref", None)
                if not ref:
                    # Ensure we still return temp_path so caller can cleanup
                    raise ValueError(f"La tabla '{table_name}' no tiene rango definido (ref).")
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                return wb, ws, ref, min_col, min_row, max_col, max_row, temp_path

    # Si no encontró tabla, cerrar wb y eliminar temp si existe antes de lanzar
    try:
        wb.close()
    except Exception:
        pass
    if temp_path:
        try:
            os.remove(temp_path)
        except Exception:
            pass

    raise ValueError(f"No se encontró la tabla '{table_name}' en el libro: {xlsx_path}")


def extraer_placas_desde_tabla(xlsx_path: str,
                               table_name: str,
                               start_row_in_table: int,
                               end_row_in_table: int,
                               target_column_index: int = 3) -> List[str]:
    """
    Devuelve una lista de placas (strings) desde la columna target de la tabla, considerando filas de datos (sin cabecera).
    """
    wb, ws, ref, min_col, min_row, max_col, max_row, temp_path = encontrar_tabla_en_libro(xlsx_path, table_name)

    # Validaciones
    cols_count = max_col - min_col + 1
    rows_total_including_header = max_row - min_row + 1
    rows_data = rows_total_including_header - 1  # quita la cabecera

    if target_column_index < 1 or target_column_index > cols_count:
        raise IndexError(f"La columna {target_column_index} está fuera del rango de la tabla (1..{cols_count}).")

    if start_row_in_table < 1 or end_row_in_table < 1:
        raise IndexError("Las filas dentro de la tabla deben ser >= 1 (sin contar cabecera).")
    if end_row_in_table < start_row_in_table:
        raise IndexError("La fila final no puede ser menor que la inicial.")
    if end_row_in_table > rows_data:
        raise IndexError(
            f"La fila final ({end_row_in_table}) excede las filas de datos ({rows_data}) en la tabla '{table_name}'."
        )

    # Coordenadas absolutas en hoja
    data_start_row_abs = min_row + 1  # primera fila de datos (justo debajo de la cabecera)
    target_col_abs = min_col + target_column_index - 1

    placas = []
    for idx in range(start_row_in_table, end_row_in_table + 1):
        r_abs = data_start_row_abs + (idx - 1)
        cell = ws.cell(row=r_abs, column=target_col_abs)
        val = cell.value
        if val is None:
            print(f"ADVERTENCIA: Celda vacía en fila {idx} de la tabla, columna {target_column_index}. Se omite.")
            continue
        placas.append(str(val).strip())

    wb.close()
    # Eliminar copia temporal si se creó
    try:
        if 'temp_path' in locals() and temp_path:
            os.remove(temp_path)
    except Exception:
        pass
    return placas


# ====== UTILIDADES DE ENTRADA/TECLAS ======
def pegar_texto_desde_clipboard(texto: str):
    """Copia al portapapeles y pega con Ctrl+V."""
    pyperclip.copy(texto)
    time.sleep(DELAY_CORTO)
    pag.hotkey("ctrl", "v")
    time.sleep(DELAY_CORTO)


def esperar_confirmacion_usuario(key: str = KEY_CONTINUAR):
    """Espera a que el usuario confirme (F8 por defecto). Si 'keyboard' no está, usa Enter."""
    mensaje = (
        f"\n🟡 Selecciona MANUALMENTE el nombre del conductor en la ventana remota.\n"
        f"Cuando hayas terminado, presiona '{key.upper()}' para continuar"
        f"{' (o Enter si keyboard no está instalado)' if not KEYBOARD_AVAILABLE else ''}..."
    )
    print(mensaje)
    if KEYBOARD_AVAILABLE:
        try:
            # Espera F8 sin bloquear otros hilos
            kb.wait(key)
        except Exception:
            input("keyboard falló; presiona Enter para continuar...")
    else:
        input("Presiona Enter para continuar...")


# ====== UTILIDADES DE CONEXIÓN Y FOCO (pywinauto) ======
def _match_sdc_title(title: str) -> bool:
    """Identifica la ventana del RemoteApp por substrings tolerantes.

    Requiere que el título contenga *ambos* 'unicon' y 'almacen' (case-insensitive).
    """
    return bool(title) and ("unicon" in title.lower()) and ("almacen" in title.lower())


def conectar_sdc():
    """Conecta a la ventana principal del SDC (ALMACEN) probando UIA/Win32.
    Devuelve (app, win) o lanza RuntimeError si no encuentra la ventana.
    """
    for backend in ("uia", "win32"):
        try:
            desktop = Desktop(backend=backend)
            candidates = [w for w in desktop.windows() if _match_sdc_title(w.window_text())]
            if candidates:
                target = candidates[0]
                app = Application(backend=backend).connect(handle=target.handle)
                win = app.window(handle=target.handle)
                print(f"[INFO] Conectado a: '{target.window_text()}' (backend={backend}, handle={target.handle})")
                return app, win
        except Exception:
            continue
    raise RuntimeError("No pude conectar al SDC (ALMACEN). Verifica que esté visible y maximizado.")


user32 = ctypes.windll.user32

def get_foreground_handle():
    return user32.GetForegroundWindow()


def is_sdc_foreground(win) -> bool:
    try:
        return get_foreground_handle() == win.handle
    except Exception:
        return False


def focus_window_hard_enter(win, retries=4, pause=0.2):
    """Trae la ventana al frente de forma 'hard' basada en ENTER (sin clic):
    - restore
    - maximize
    - set_focus
    - enviar ENTER para afianzar el foco (simula interacción)
    Retorna True si logra poner SDC en foreground; False si no.
    """
    for i in range(retries):
        try:
            win.restore()
        except Exception:
            pass
        try:
            win.maximize()
        except Exception:
            pass
        try:
            win.set_focus()
            time.sleep(pause)
        except Exception:
            pass
        # Empujón de interacción: ENTER (sin clic)
        try:
            send_keys("{ENTER}")
            time.sleep(pause)
        except Exception:
            pass

        if is_sdc_foreground(win):
            return True
        time.sleep(pause)
    return False


def go_to_sdc(win, attempts: int = 5, timeout=6.0) -> bool:
    """Intenta llevar el foco al `win` del SDC de forma robusta.
    Combina llamadas directas a pywinauto con pulsaciones de Alt+Tab como respaldo.
    """
    t0 = time.time()
    time.sleep(1.0)  # espera inicial antes de intentar 

    # ALT+TAB gradual: en el intento n enviamos ALT+TAB n veces (manteniendo ALT)
    for attempt in range(1, attempts + 1):
        print(f"[INFO] Intento {attempt} traer SDC al frente con ALT+TAB x{attempt}...")
        # Construir tecla que mantiene ALT y pulsa TAB `attempt` veces: "%({TAB}{TAB}...)"
        tabs = "".join(["{TAB}"] * attempt)
        keys = f"%({tabs})"
        try:
            send_keys(keys)
        except Exception:
            # fallback a un solo Alt+Tab si la construcción falla
            send_keys("%{TAB}")

        time.sleep(0.25)
        if focus_window_hard_enter(win):
            return True
        if time.time() - t0 > timeout:
            break

# Tiempo máximo a esperar a que cierre la ventana "Registro de Salidas por Venta - \\Remota"
DEFAULT_REGISTRO_TIMEOUT = 20.0  # segundos
POLL_INTERVAL = 0.5

def wait_for_registro_salidas_close(title_substr: str = r"Registro de Salidas por Venta - \\Remota",
                                    timeout: float = DEFAULT_REGISTRO_TIMEOUT,
                                    poll_interval: float = POLL_INTERVAL) -> bool:
    """
    Espera hasta que no exista una ventana cuyo título contenga `title_substr`.
    Usa backend 'win32' ya que la ventana sólo es visible con ese backend.
    Retorna True si desaparece antes del timeout, False si vence el timeout o ocurre un fallo.
    """
    try:
        desktop = Desktop(backend="win32")
    except Exception:
        # No podemos usar win32: no bloquear el flujo
        return True

    t0 = time.time()
    while True:
        try:
            found = any(title_substr.lower() in w.window_text().lower() for w in desktop.windows())
        except Exception:
            # En caso de error al listar ventanas, continuar para evitar bloqueo
            return True

        if not found:
            return True
        if time.time() - t0 > timeout:
            print(f"[WARN] Timeout ({timeout}s) esperando que desaparezca '{title_substr}'")
            return False
        time.sleep(poll_interval)

def sdc_active(win, poll_interval: float = 0.25) -> bool:
    """
    Bloquea hasta que la ventana `win` del SDC esté en foreground y su título coincida con _match_sdc_title().
    Retorna True cuando la ventana está activa.
    """
    while True:
        try:
            # Si el foreground coincide con el handle y el título coincide, consideramos la ventana activa
            if is_sdc_foreground(win):
                try:
                    title = win.window_text()
                except Exception:
                    title = ""
                if _match_sdc_title(title):
                    return True            
        except KeyboardInterrupt:
            raise
        except Exception:
            # Ignorar errores temporales y seguir intentando
            pass
        time.sleep(poll_interval)

# ====== FLUJO DE DESPACHO EN REMOTO ======
def wait_for_informacion_window(title_substr: str = r"Información - \\Remota",
                                timeout: float = 10.0,
                                poll_interval: float = 0.5) -> bool:
    """
    Espera hasta que aparezca una ventana cuyo título contenga `title_substr`.
    Una vez que aparece, envía SPACE para cerrarla y continuar.
    Retorna True si la ventana apareció y se cerró, False si vence el timeout.
    """
    try:
        desktop = Desktop(backend="win32")
    except Exception:
        print("[WARN] No se pudo usar win32 backend para esperar ventana Información")
        return False

    t0 = time.time()
    while True:
        try:
            windows = desktop.windows()
            found = any(title_substr.lower() in w.window_text().lower() for w in windows)
        except Exception:
            found = False

        if found:
            print(f"[INFO] Ventana '{title_substr}' detectada. Enviando SPACE para cerrar...")
            time.sleep(0.5)  # pequeña pausa para asegurar que la ventana esté lista
            send_keys("{SPACE}")
            time.sleep(DELAY_MEDIO)
            return True
        
        if time.time() - t0 > timeout:
            print(f"[WARN] Timeout ({timeout}s) esperando que aparezca '{title_substr}'")
            return False
        
        time.sleep(poll_interval)


def flujo_despacho_para_placa(placa: str):
    """
    Ejecuta la secuencia de teclas en la ventana remota para procesar una placa.
    Se asume que la ventana remota ya está en foco y en estado inicial.
    """
    print(f"\n➡️ Procesando placa: {placa}")

    # 2) Tecla D (botón 'Despacho')
    pag.press("d")
    time.sleep(DELAY_LARGO)

    # 3) 8× Shift+Tab → Espacio → 'alp' → Tab → Espacio → 'A'
    for _ in range(SHIFT_TABS_A_BOTON_NOMBRE):
        send_keys('+{TAB}')
    time.sleep(DELAY_CORTO)

    send_keys("{SPACE}")  # abre selector de nombre
    time.sleep(DELAY_MEDIO)

    pag.typewrite(FILTRO_NOMBRE_TEXTO, interval=0.05)  # escribe "alp"
    time.sleep(DELAY_MEDIO)

    send_keys("{TAB}")
    time.sleep(DELAY_CORTO)

    send_keys("{SPACE}")  # confirma selección (abre/acepta según UI)
    time.sleep(DELAY_MEDIO)

    pag.press("a")  # Aceptar (cierra ventana de nombre)
    time.sleep(DELAY_MEDIO)

    # 4) Tab → Espacio (abre ventana donde se ingresa la placa)
    send_keys("{TAB}")
    time.sleep(DELAY_CORTO)
    send_keys("{SPACE}")
    time.sleep(DELAY_MEDIO)

    # 5) Shift+Tab → Ctrl+V para pegar placa
    send_keys('+{TAB}')
    time.sleep(DELAY_CORTO)
    pegar_texto_desde_clipboard(placa)
    time.sleep(DELAY_MEDIO)

    # 6) Shift+Tab → Espacio (acepta)
    send_keys('+{TAB}')
    time.sleep(DELAY_CORTO)
    send_keys("{SPACE}")
    time.sleep(DELAY_MEDIO)

    # 6.1) Shift+Tab)x2
    send_keys('+{TAB}')
    send_keys('+{TAB}')

    # 7) Pausa para selección manual del conductor → F8 para continuar
    esperar_confirmacion_usuario(KEY_CONTINUAR)

    # 8) A → A para cerrar ventanas
    pag.press("a")
    time.sleep(DELAY_MEDIO)
    pag.press("a") # c Cerrar para pruebas
    time.sleep(DELAY_LARGO)
    send_keys("{SPACE}")
    time.sleep(DELAY_MEDIO)

    # Esperar para que aparezca la pantalla "Información - \\Remota" → SPACE para continuar
    wait_for_informacion_window(r"Información - \\Remota", timeout=100.0)

    # Esperar que la ventana Registro de Salidas por Venta - \\Remota se cierre (win32 only)
    """ wait_for_registro_salidas_close(r"Registro de Salidas por Venta - \\Remota", timeout=DEFAULT_REGISTRO_TIMEOUT) """
    
    # Intentar conectar y enfocar SDC automáticamente
    """ try:
        app, win = conectar_sdc()
    except Exception as e:
        print(f"[ERROR] No pude conectar al SDC: {e}")
        sys.exit(1)

    if sdc_active(win):
        print("[INFO] SDC es la ventana activa.") """

    # 7) Pausa para selección manual del conductor → F8 para continuar
    """ esperar_confirmacion_usuario(KEY_CONTINUAR) """

# ============== FLUJO PRINCIPAL ==============
def main():
    print("Cargando placas desde Excel...")
    try:
        placas = extraer_placas_desde_tabla(EXCEL_PATH,
                                            TABLE_NAME,
                                            START_ROW_IN_TABLE,
                                            END_ROW_IN_TABLE,
                                            TARGET_COLUMN_INDEX)
    except Exception as e:
        print(f"\n❌ Error leyendo Excel/Tabla: {e}")
        sys.exit(1)

    if not placas:
        print("No se encontraron placas en el rango especificado.")
        sys.exit(0)

    print(f"✅ {len(placas)} placa(s) lista(s): {placas}")

    # Intentar conectar y enfocar SDC automáticamente
    try:
        app, win = conectar_sdc()
    except Exception as e:
        print(f"[ERROR] No pude conectar al SDC: {e}")
        sys.exit(1)

    if not go_to_sdc(win):
        print("[WARN] No pude recuperar foco del SDC tras leer Excel.")
    else:
        print("[INFO] Ventana SDC enfocada.")
        

    # Itera placas
    # 1) Enter para cargar datos
    pag.press("enter")
    time.sleep(DELAY_MEDIO)

    for placa in placas:
        flujo_despacho_para_placa(placa)

    print("\n✅ Proceso completado para todas las placas.")


if __name__ == "__main__":
    main()
# Fin del script despacho_placas.py